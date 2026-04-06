import base64
import io
import json
import os
import urllib.parse
import zipfile
from datetime import datetime
from typing import Any

import aiohttp
import openpyxl
import pandas as pd
from fastmcp import FastMCP
from openai import AsyncOpenAI
from openpyxl.styles import PatternFill
from tqdm.asyncio import tqdm

from config import MCPSettings

# Initialize MCP
mcp = FastMCP("Site Check Pipeline")

# Load settings locally
settings = MCPSettings()  # type: ignore

# Initialize OpenAI client for OpenRouter
openai_client = AsyncOpenAI(
    base_url="https://openrouter.ai/api/v1",
    api_key=settings.openrouter_api_key.get_secret_value(),
)


async def check_street_view_metadata(
    session: aiohttp.ClientSession, address: str
) -> dict[str, Any]:
    url = "https://maps.googleapis.com/maps/api/streetview/metadata"
    params = {"location": address, "key": settings.gcp_api_key.get_secret_value()}
    async with session.get(url, params=params) as response:
        return await response.json()


async def fetch_street_view_image(
    session: aiohttp.ClientSession, address: str
) -> bytes | None:
    url = "https://maps.googleapis.com/maps/api/streetview"
    params = {
        "size": "600x600",
        "location": address,
        "key": settings.gcp_api_key.get_secret_value(),
        "source": "outdoor",
        "return_error_code": "true",
    }
    async with session.get(url, params=params) as response:
        if response.status == 200:
            return await response.read()
        return None


async def analyze_image_with_vision_model(
    image_bytes: bytes, address: str
) -> dict[str, Any]:
    base64_image = base64.b64encode(image_bytes).decode("utf-8")
    prompt = """Please analyze the provided Google Street View image of the building at the location.
Extract the following information in a structured JSON format:
- "unit_state": string (MUST be exactly one of: "no visible damage", "slight damage", "heavy damage", "not detected")
- "analyst_notes": string (Brief summary of overall impression and any visible details)

Return ONLY valid JSON with exactly these keys."""

    try:
        response = await openai_client.chat.completions.create(
            model=settings.vision_model,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/jpeg;base64,{base64_image}"
                            },
                        },
                    ],
                }
            ],
            response_format={"type": "json_object"},
        )

        content = response.choices[0].message.content
        if content is None:
            return {"error": "Vision analysis returned empty string"}
        return json.loads(content)
    except Exception as e:
        print(f"\n⚠️ WARNING: Vision Analysis API failed for '{address}': {e}")
        return {"error": f"Failed Vision Analysis: {e}"}


async def process_single_address(
    session: aiohttp.ClientSession, address: str, dry_run: bool
) -> dict[str, Any]:
    encoded_address = urllib.parse.quote(address)
    map_url = f"https://www.google.com/maps/search/?api=1&query={encoded_address}"

    result = {
        "Address": address,
        "Status": None,
        "Google_Maps_Link": map_url,
        "Image_Date": None,
        "Street_View_Link": None,
        "Unit_State": None,
        "Notes": None,
        "Error": None,
        "_image_bytes": None,
    }

    # Check Metadata
    try:
        metadata = await check_street_view_metadata(session, address)
        status = metadata.get("status")
        result["Status"] = status
        if status == "OK":
            result["Image_Date"] = metadata.get("date")
            # Create a precise keyless Street View link using the Pano ID
            pano_id = metadata.get("pano_id")
            if pano_id:
                result["Street_View_Link"] = f"https://www.google.com/maps/@?api=1&map_action=pano&pano={pano_id}"
            else:
                # Fallback to viewpoint if pano_id is missing
                location = metadata.get("location")
                if location:
                    lat, lng = location.get("lat"), location.get("lng")
                    if lat and lng:
                        result["Street_View_Link"] = f"https://www.google.com/maps/@?api=1&map_action=pano&viewpoint={lat},{lng}"
        else:
            error_msg = metadata.get("error_message", "")
            print(
                f"\n⚠️ WARNING: Google API returned {status} for '{address}'. {error_msg}"
            )

        if dry_run or status != "OK":
            return result
    except Exception as e:
        print(f"\n⚠️ ERROR: Network failure checking metadata for '{address}': {e}")
        result["Error"] = f"Metadata check failed: {e}"
        return result

    # Fetch Image
    try:
        image_bytes = await fetch_street_view_image(session, address)
        if not image_bytes:
            print(
                f"\n⚠️ WARNING: Failed to fetch the actual image bytes for '{address}' from Google Maps."
            )
            result["Error"] = "Failed to fetch image"
            return result
        result["_image_bytes"] = image_bytes
    except Exception as e:
        print(f"\n⚠️ ERROR: Network failure fetching image for '{address}': {e}")
        result["Error"] = f"Image fetch failed: {e}"
        return result

    # Vision Analysis
    try:
        analysis = await analyze_image_with_vision_model(image_bytes, address)
        if "error" in analysis:
            result["Error"] = analysis["error"]
        else:
            result["Unit_State"] = analysis.get("unit_state", "not detected")
            result["Notes"] = analysis.get("analyst_notes")
    except Exception as e:
        result["Error"] = f"Vision analysis failed: {e}"

    return result


@mcp.tool()
async def process_locations_batch(
    addresses: list[str], dry_run: bool = False, progress_callback: Any = None
) -> str:
    """
    Site Check Pipeline - Processes a batch of structured addresses.

    Args:
        addresses: A clean, structured list of address strings to process.
        dry_run: If True, only checks metadata to save Vision model costs.
        progress_callback: Optional async function to report progress.

    Returns:
        A machine-readable JSON string containing completion status and report paths.
    """
    print(
        f"Starting batch processing of {len(addresses)} addresses (Dry run: {dry_run})..."
    )

    results = []
    async with aiohttp.ClientSession() as session:
        tasks = [
            process_single_address(session, address, dry_run) for address in addresses
        ]

        for f in tqdm.as_completed(
            tasks, total=len(addresses), desc="Processing Addresses"
        ):
            res = await f
            results.append(res)
            if progress_callback:
                try:
                    # Report progress back to the caller
                    await progress_callback(
                        f"Audit Progress: {len(results)}/{len(addresses)} sites processed ({res.get('Address', 'Unknown')})"
                    )
                except Exception:
                    pass

    # Sort results to maintain a consistent order (optional, but good for indexing)


    for res in results:
        # Cleanup temporary image bytes before creating DataFrame
        if "_image_bytes" in res:
            del res["_image_bytes"]

    df = pd.DataFrame(results)

    now = datetime.now()
    age_strings = []
    age_months_list = []

    for d_str in df.get("Image_Date", []):
        if not d_str or pd.isna(d_str):
            age_strings.append("Unknown")
            age_months_list.append(None)
            continue
        try:
            d = datetime.strptime(str(d_str), "%Y-%m")
            m_diff = (now.year - d.year) * 12 + (now.month - d.month)
            age_months_list.append(m_diff)

            if m_diff < 12:
                age_strings.append(f"{m_diff} month{'s' if m_diff != 1 else ''}")
            else:
                yrs = m_diff // 12
                mos = m_diff % 12
                s = f"{yrs} yr{'s' if yrs != 1 else ''}"
                if mos > 0:
                    s += f" {mos} mo."
                age_strings.append(s)
        except Exception:
            age_strings.append("Unknown")
            age_months_list.append(None)

    if "Image_Date" in df.columns:
        df.insert(df.columns.get_loc("Image_Date") + 1, "Image_Age", age_strings)
    else:
        df["Image_Age"] = age_strings

    df["Image_Age_Months"] = age_months_list

    output_file = "data/site_check_report.xlsx"
    os.makedirs("data", exist_ok=True)
    df.to_excel(output_file, index=False)

    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Column name mapping to find indices
    col_names = [cell.value for cell in ws[1]] if ws else []
    
    # 1. Format Hyperlinks (Google Maps, Street View)
    hyperlink_cols = ["Google_Maps_Link", "Street_View_Link"]
    
    for col_name in hyperlink_cols:
        if col_name in col_names:
            col_idx = col_names.index(col_name) + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    url = cell.value
                    cell.value = f'=HYPERLINK("{url}", "Click to View")'
                    cell.style = "Hyperlink"

    # 2. Apply Aging Colors (Image_Age)
    if "Image_Age" in col_names:
        age_col_idx = col_names.index("Image_Age") + 1
        months_col_idx = col_names.index("Image_Age_Months") + 1 if "Image_Age_Months" in col_names else None

        for row_idx in range(2, ws.max_row + 1):
            if months_col_idx:
                months_val = df.iloc[row_idx-2]["Image_Age_Months"] # Use df for numeric value
                if pd.isna(months_val) or months_val is None:
                    continue

                m = min(int(months_val), 120)
                if m < 60:
                    r = int((m / 60) * 255)
                    g = 255
                else:
                    r = 255
                    g = int(255 - ((m - 60) / 60) * 200)

                hex_color = f"FF{r:02X}{g:02X}00"
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                ws.cell(row=row_idx, column=age_col_idx).fill = fill

    # 3. Clean up helper columns
    if "Image_Age_Months" in col_names:
        idx = col_names.index("Image_Age_Months") + 1
        ws.delete_cols(idx)

    try:
        wb.save(output_file)
    except Exception as e:
        print(f"Warning: Could not save formatted Excel file: {e}")
        df.drop(columns=["Image_Age_Months"], errors="ignore").to_excel(
            "data/site_check_report_unformatted.xlsx", index=False
        )
        output_file = "data/site_check_report_unformatted.xlsx"

    # Save machine-readable CSV version
    csv_output_file = "data/site_check_report.csv"
    df.drop(columns=["Image_Age_Months"], errors="ignore").to_csv(csv_output_file, index=False)

    summary = f"Batch pipeline complete. Processed {len(addresses)} addresses."
    if dry_run:
        summary += " (DRY RUN - metadata only)."

    result_payload = {
        "status": "completed",
        "summary": summary,
        "files": {
            "excel": os.path.abspath(output_file),
            "csv": os.path.abspath(csv_output_file)
        },
        "count": len(addresses)
    }

    return json.dumps(result_payload, indent=2)


if __name__ == "__main__":
    mcp.run()
