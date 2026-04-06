import asyncio
import json
import os
import logging
from datetime import datetime
from typing import Any
from importlib.metadata import version

import aiohttp
import openpyxl
import pandas as pd
from fastmcp import FastMCP
from openpyxl.styles import PatternFill
from fastmcp import Context
from mcp_server.utils import check_street_view_metadata, get_google_maps_link, get_street_view_link, get_cardinal_direction
from mcp_server.vision import fetch_street_view_image, analyze_image_with_vision_model
from mcp_server.settings import Settings

# Initialize Settings
settings = Settings()

# Ensure logs directory exists
from pathlib import Path
log_dir = Path("logs")
log_dir.mkdir(exist_ok=True)

# Configure Logging
logging.basicConfig(
    level=settings.log_level,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)
logger.info(f"Starting MCP Server with Log Level: {settings.log_level}")

# Get version from package metadata
VERSION = version("mcp-sitecheck")

# Initialize MCP
mcp = FastMCP("Site Check Pipeline", version=VERSION)

def flatten_analysis_result(analysis: dict[str, Any]) -> dict[str, Any]:
    """
    Flatten nested structures in the analysis result to ensure simple scalar values
    for spreadsheet columns.
    """
    flattened = {}
    for key, value in analysis.items():
        if isinstance(value, dict):
            # If it's a dict, prefer 'description' or 'value' or 'result'
            if "description" in value:
                flattened[key] = value["description"]
            elif "value" in value:
                flattened[key] = value["value"]
            elif "result" in value:
                flattened[key] = value["result"]
            else:
                # Fallback: just stringify the whole thing if we can't find a primary field
                flattened[key] = str(value)
        elif isinstance(value, list):
            # Join lists into a comma-separated string
            flattened[key] = ", ".join(str(v) for v in value)
        else:
            flattened[key] = value
    return flattened

async def process_single_address(
    session: aiohttp.ClientSession, 
    address: str, 
    dry_run: bool, 
    analysis_prompt: str, 
    analysis_schema: str,
    output_dir: str = None,
    ctx: Context | None = None
) -> dict[str, Any]:

    result = {
        "Address": address,
        "Status": None,
        "Google_Maps_Link": get_google_maps_link(address),
        "Image_Date": None,
        "Error": None,
        "_images": [],
    }

    # Check Metadata
    try:
        metadata = await check_street_view_metadata(session, address)
        status = metadata.get("status")
        result["Status"] = status
        
        # Always include date and link if available
        result["Image_Date"] = metadata.get("date")

        if status != "OK":
            error_msg = metadata.get("error_message", "No imagery available at this location.")
            result["Error"] = f"Google API: {status}. {error_msg}"
            return result
        
        # Dynamic Street View Links by Direction
        num_images = settings.street_view_image_count
        headings = [int(i * 360 / num_images) for i in range(num_images)]
        
        for heading in headings:
            cardinal = get_cardinal_direction(heading)
            col_name = f"Street_View_{cardinal}_{heading}deg"
            result[col_name] = get_street_view_link(metadata, heading=heading)

        if dry_run:
            return result
    except Exception as e:
        result["Error"] = f"Metadata check failed: {e}"
        return result

    # Fetch Images (360 view) in parallel
    try:
        image_tasks = [fetch_street_view_image(session, address, heading=h) for h in headings]
        image_responses = await asyncio.gather(*image_tasks)
        
        for heading, image_bytes in zip(headings, image_responses):
            if image_bytes:
                result["_images"].append({"bytes": image_bytes, "heading": heading})
        
        if not result["_images"]:
            result["Error"] = "Failed to fetch any image bytes"
            return result
    except Exception as e:
        result["Error"] = f"Image fetch failed: {e}"
        return result

    # Vision Analysis
    try:
        analysis = await analyze_image_with_vision_model(result["_images"], address, analysis_prompt, analysis_schema, output_dir=output_dir)
        if "error" in analysis:
            result["Error"] = analysis["error"]
            result["_validation_error"] = True # Treat vision errors as validation issues for sorting
        else:
            # Capture validation error flag
            result["_validation_error"] = analysis.get("_validation_error", False)
            # Flatten analysis before updating result
            result.update(flatten_analysis_result(analysis))
    except Exception as e:
        result["Error"] = f"Vision analysis failed: {e}"
        result["_validation_error"] = True

    return result

@mcp.tool()
async def process_locations_batch(
    addresses: list[str] = None, 
    source_file: str = None,
    filter_query: str = None,
    address_column: str = None,
    analysis_prompt: str = None,
    analysis_schema: str = None,
    dry_run: bool = False, 
    output_dir: str = "output",
    timeout: int = None,
    ctx: Context = None
) -> str:
    """
    Site Check Pipeline - Processes a batch of structured addresses with 360-degree vision audit.
    
    Can take a direct list of addresses OR a source_file (CSV/JSON) with an optional filter_query (pandas style).
    If using source_file, you MUST specify address_column.

    Args:
        addresses: A clean, structured list of address strings to process.
        source_file: Optional: Path to a CSV or JSON file containing addresses.
        filter_query: Optional: Pandas-style query string to filter the source_file (e.g., "State == 'CA'").
        address_column: MANDATORY if using source_file: The column name in the source_file that contains the addresses.
        analysis_prompt: MANDATORY: A detailed description of what to look for in the images.
        analysis_schema: MANDATORY: JSON schema string for structured generation.
        dry_run: If True, only checks metadata to save Vision model costs.
        output_dir: The directory where the results will be saved.
        timeout: Optional: Maximum time in seconds for the entire batch. Defaults to DEFAULT_TIMEOUT in .env.
        ctx: MCP Context for progress reporting.
    """
    if not analysis_prompt or not analysis_schema:
        return "Error: analysis_prompt and analysis_schema are mandatory."

    # Use default timeout from settings if not provided
    effective_timeout = timeout if timeout is not None else settings.default_timeout

    final_addresses = addresses or []
    
    # Resolve source_file if provided
    if source_file:
        if not address_column:
            return "Error: address_column must be specified when using source_file."
        try:
            if source_file.endswith('.csv'):
                # Try latin-1 first as many property datasets use it
                try:
                    df_source = pd.read_csv(source_file, encoding='latin-1')
                except Exception:
                    df_source = pd.read_csv(source_file)
            elif source_file.endswith('.json'):
                df_source = pd.read_json(source_file)
            else:
                return f"Error: Unsupported file format for {source_file}. Use CSV or JSON."

            if filter_query:
                df_source = df_source.query(filter_query)
            
            if address_column not in df_source.columns:
                return f"Error: Column '{address_column}' not found in {source_file}. Available: {list(df_source.columns)}"
            
            file_addresses = df_source[address_column].dropna().unique().tolist()
            final_addresses.extend([str(a).strip() for a in file_addresses])
        except Exception as e:
            return f"Error reading source_file: {str(e)}"

    if not final_addresses:
        return "Error: No addresses provided or found in source_file."

    results = []
    total = len(final_addresses)
    completed = 0
    
    # Process batch in parallel without restrictions for maximum speed
    async def wrapped_process(session, address):
        nonlocal completed
        res = await process_single_address(session, address, dry_run, analysis_prompt, analysis_schema, output_dir, ctx)
        completed += 1
        if ctx:
            await ctx.report_progress(completed, total)
        return res

    async with aiohttp.ClientSession() as session:
        tasks = [asyncio.create_task(wrapped_process(session, addr)) for addr in final_addresses]
        
        if effective_timeout:
            done, pending = await asyncio.wait(tasks, timeout=effective_timeout)
            
            # Cancel pending tasks
            for task in pending:
                task.cancel()
            
            # Get results from done tasks
            results = [task.result() for task in done]
            
            # Mark timed out addresses
            finished_addresses = {res["Address"] for res in results}
            for addr in final_addresses:
                if addr not in finished_addresses:
                    results.append({
                        "Address": addr,
                        "Status": "Timeout",
                        "Error": f"Operation exceeded time cap of {effective_timeout}s",
                        "_validation_error": True
                    })
        else:
            results = await asyncio.gather(*tasks)

    # Ensure all results have the _validation_error key and vision columns
    vision_keys = []
    try:
        schema_dict = json.loads(analysis_schema)
        vision_keys = list(schema_dict.get("properties", {}).keys())
    except Exception:
        pass

    for res in results:
        # Determine if it's a validation/error case
        if "_validation_error" not in res:
            res["_validation_error"] = True if res.get("Error") else False
        
        # Ensure all vision columns exist to avoid empty/NaN cells in Excel
        for key in vision_keys:
            if key not in res:
                res[key] = "N/A"

        if "_images" in res:
            del res["_images"]

    # Sort results: Valid results (False) first, validation errors/errors (True) at the bottom
    results.sort(key=lambda x: x.get("_validation_error", False))

    df = pd.DataFrame(results)
    
    # Calculate Image Age
    now = datetime.now()
    age_strings = []
    age_months_list = []

    for d_str in df.get("Image_Date", []):
        if not d_str or pd.isna(d_str):
            age_strings.append("Unknown")
            age_months_list.append(None)
            continue
        try:
            d = datetime.strptime(str(d_str), "%Y-%m")
            m_diff = (now.year - d.year) * 12 + (now.month - d.month)
            age_months_list.append(m_diff)

            if m_diff < 12:
                age_strings.append(f"{m_diff} month{'s' if m_diff != 1 else ''}")
            else:
                yrs = m_diff // 12
                mos = m_diff % 12
                s = f"{yrs} yr{'s' if yrs != 1 else ''}"
                if mos > 0:
                    s += f" {mos} mo."
                age_strings.append(s)
        except Exception:
            age_strings.append("Unknown")
            age_months_list.append(None)

    if "Image_Date" in df.columns:
        df.insert(df.columns.get_loc("Image_Date") + 1, "Image_Age", age_strings)
    else:
        df["Image_Age"] = age_strings
    df["Image_Age_Months"] = age_months_list

    # Reorder columns: Address, [MCP/Vision Cols], Status, Image_Date, Image_Age, Error, Google_Maps_Link, [Street_View_*]
    cols = list(df.columns)
    
    # Define groups
    head = ["Address"]
    metadata = ["Status", "Image_Date", "Image_Age", "Error"]
    links_primary = ["Google_Maps_Link"]
    links_sv = [c for c in cols if c.startswith("Street_View_")]
    
    # Known fixed columns to exclude from "vision/mcp" group
    fixed = head + metadata + links_primary + links_sv + ["Image_Age_Months"]
    
    # Everything else is a vision/mcp analysis column
    vision_cols = [c for c in cols if c not in fixed]
    
    # Final desired order
    ordered_cols = head + vision_cols + [m for m in metadata if m in cols] + links_primary + links_sv + (["Image_Age_Months"] if "Image_Age_Months" in cols else [])
    
    # Filter to only existing columns
    ordered_cols = [c for c in ordered_cols if c in df.columns]
    df = df[ordered_cols]
    
    # Drop internal helper column before saving
    df = df.drop(columns=["_validation_error"], errors="ignore")

    # Save outputs
    from pathlib import Path
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    excel_file = os.path.join(output_dir, "site_check_report.xlsx")
    jsonl_file = os.path.join(output_dir, "site_check_report.jsonl")
    
    df.to_excel(excel_file, index=False)
    
    # Post-process Excel formatting
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        col_names = [cell.value for cell in ws[1]] if ws else []
        
        # Format Hyperlinks
        for col_name in col_names:
            if col_name == "Google_Maps_Link" or col_name.startswith("Street_View_"):
                col_idx = col_names.index(col_name) + 1
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        url = str(cell.value)
                        if url and not url.startswith('=HYPERLINK'):
                            cell.hyperlink = url
                            cell.value = url
                        cell.style = "Hyperlink"

        # Apply Aging Colors (Row-wide)
        if "Image_Age_Months" in col_names:
            months_col_idx = col_names.index("Image_Age_Months") + 1
            for row_idx in range(2, ws.max_row + 1):
                months_val = ws.cell(row=row_idx, column=months_col_idx).value
                if months_val is not None:
                    m = int(months_val)
                    if m >= 12:
                        # 1 year or older: Red
                        hex_color = "FFFF0000"
                    else:
                        # 0-11 months: Gradual Green to Red
                        # 0 months -> Green (00FF00)
                        # 11 months -> Almost Red
                        r = int((m / 11) * 255)
                        g = int(255 - (m / 11) * 255)
                        hex_color = f"FF{r:02X}{g:02X}00"
                    
                    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    for col_idx in range(1, len(col_names) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

        # Delete helper column
        if "Image_Age_Months" in col_names:
            ws.delete_cols(col_names.index("Image_Age_Months") + 1)
        
        wb.save(excel_file)
    except Exception as e:
        print(f"Excel formatting failed: {e}")

    df.drop(columns=["Image_Age_Months"], errors="ignore").to_json(jsonl_file, orient="records", lines=True)

    return json.dumps({
        "status": "completed",
        "count": len(final_addresses),
        "files": {
            "excel": excel_file,
            "jsonl": jsonl_file
        }
    }, indent=2)

if __name__ == "__main__":
    mcp.run()
