import json
import os
from datetime import datetime
from typing import Any

import aiohttp
import openpyxl
import pandas as pd
from fastmcp import FastMCP
from openpyxl.styles import PatternFill
from tqdm.asyncio import tqdm

from mcp_server.utils import check_street_view_metadata, get_google_maps_link, get_street_view_link
from mcp_server.vision import fetch_street_view_image, analyze_image_with_vision_model

# Initialize MCP
mcp = FastMCP("Site Check Pipeline")

async def process_single_address(
    session: aiohttp.ClientSession, 
    address: str, 
    dry_run: bool, 
    analysis_prompt: str, 
    analysis_schema: str
) -> dict[str, Any]:
    result = {
        "Address": address,
        "Status": None,
        "Google_Maps_Link": get_google_maps_link(address),
        "Image_Date": None,
        "Street_View_Link": None,
        "Error": None,
        "_image_bytes": None,
    }

    # Check Metadata
    try:
        metadata = await check_street_view_metadata(session, address)
        status = metadata.get("status")
        result["Status"] = status
        
        # Always include date and link if available
        result["Image_Date"] = metadata.get("date")
        result["Street_View_Link"] = get_street_view_link(metadata)

        if status != "OK":
            error_msg = metadata.get("error_message", "No imagery available at this location.")
            result["Error"] = f"Google API: {status}. {error_msg}"
            return result

        if dry_run:
            return result
    except Exception as e:
        result["Error"] = f"Metadata check failed: {e}"
        return result

    # Fetch Image
    try:
        image_bytes = await fetch_street_view_image(session, address)
        if not image_bytes:
            result["Error"] = "Failed to fetch image bytes"
            return result
        result["_image_bytes"] = image_bytes
    except Exception as e:
        result["Error"] = f"Image fetch failed: {e}"
        return result

    # Vision Analysis
    try:
        analysis = await analyze_image_with_vision_model(image_bytes, address, analysis_prompt, analysis_schema)
        if "error" in analysis:
            result["Error"] = analysis["error"]
        else:
            result.update(analysis)
    except Exception as e:
        result["Error"] = f"Vision analysis failed: {e}"

    return result

@mcp.tool()
async def process_locations_batch(
    addresses: list[str], 
    analysis_prompt: str,
    analysis_schema: str,
    dry_run: bool = False, 
) -> str:
    """
    Site Check Pipeline - Processes a batch of structured addresses.

    Args:
        addresses: A clean, structured list of address strings to process.
        analysis_prompt: MANDATORY: Custom prompt for the vision model checking the image.
        analysis_schema: MANDATORY: JSON schema string for structured generation.
        dry_run: If True, only checks metadata to save Vision model costs.

    Returns:
        A machine-readable JSON string containing completion status and report paths.
    """
    results = []
    async with aiohttp.ClientSession() as session:
        tasks = [
            process_single_address(session, address, dry_run, analysis_prompt, analysis_schema) for address in addresses
        ]

        for f in tqdm.as_completed(tasks, total=len(addresses), desc="Processing Addresses"):
            res = await f
            results.append(res)

    for res in results:
        if "_image_bytes" in res:
            del res["_image_bytes"]

    df = pd.DataFrame(results)
    
    # Calculate Image Age
    now = datetime.now()
    age_strings = []
    age_months_list = []

    for d_str in df.get("Image_Date", []):
        if not d_str or pd.isna(d_str):
            age_strings.append("Unknown")
            age_months_list.append(None)
            continue
        try:
            d = datetime.strptime(str(d_str), "%Y-%m")
            m_diff = (now.year - d.year) * 12 + (now.month - d.month)
            age_months_list.append(m_diff)

            if m_diff < 12:
                age_strings.append(f"{m_diff} month{'s' if m_diff != 1 else ''}")
            else:
                yrs = m_diff // 12
                mos = m_diff % 12
                s = f"{yrs} yr{'s' if yrs != 1 else ''}"
                if mos > 0:
                    s += f" {mos} mo."
                age_strings.append(s)
        except Exception:
            age_strings.append("Unknown")
            age_months_list.append(None)

    if "Image_Date" in df.columns:
        df.insert(df.columns.get_loc("Image_Date") + 1, "Image_Age", age_strings)
    else:
        df["Image_Age"] = age_strings
    df["Image_Age_Months"] = age_months_list

    # Save outputs
    os.makedirs("data", exist_ok=True)
    excel_file = "data/site_check_report.xlsx"
    csv_file = "data/site_check_report.csv"
    
    df.to_excel(excel_file, index=False)
    
    # Post-process Excel formatting
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        col_names = [cell.value for cell in ws[1]] if ws else []
        
        # Format Hyperlinks
        for col_name in ["Google_Maps_Link", "Street_View_Link"]:
            if col_name in col_names:
                col_idx = col_names.index(col_name) + 1
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell.value = f'=HYPERLINK("{cell.value}", "Click to View")'
                        cell.style = "Hyperlink"

        # Apply Aging Colors
        if "Image_Age" in col_names and "Image_Age_Months" in col_names:
            age_col_idx = col_names.index("Image_Age") + 1
            months_col_idx = col_names.index("Image_Age_Months") + 1
            for row_idx in range(2, ws.max_row + 1):
                months_val = ws.cell(row=row_idx, column=months_col_idx).value
                if months_val is not None:
                    m = min(int(months_val), 120)
                    r = int((m / 60) * 255) if m < 60 else 255
                    g = 255 if m < 60 else int(255 - ((m - 60) / 60) * 200)
                    hex_color = f"FF{r:02X}{g:02X}00"
                    ws.cell(row=row_idx, column=age_col_idx).fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        # Delete helper column
        if "Image_Age_Months" in col_names:
            ws.delete_cols(col_names.index("Image_Age_Months") + 1)
        
        wb.save(excel_file)
    except Exception as e:
        print(f"Excel formatting failed: {e}")

    df.drop(columns=["Image_Age_Months"], errors="ignore").to_csv(csv_file, index=False)

    return json.dumps({
        "status": "completed",
        "count": len(addresses),
        "files": {
            "excel": os.path.abspath(excel_file),
            "csv": os.path.abspath(csv_file)
        }
    }, indent=2)

if __name__ == "__main__":
    mcp.run()
